import telebot
from telebot import types
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from collections import defaultdict
import Panel_Main
import Panel_Video
import Panel_Link
import Panel_Table
import Panel_Addition
import Panel_IOS
import Panel_Android

class TG_BOT:
    def __init__(self):
        self.bot = telebot.TeleBot('token')
        yan = 14                #Курс юаня
        liberty = 7             #Процент выкупа
        delivery_price = 620    #Цена за доставку в кг
        self.cell = defaultdict(list)

        font = Font(
        name='Times New Roman',
        size=14,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
            )
        alignment = Alignment(
                horizontal='center',
                vertical='center',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0
                   )

        thing_row = defaultdict(list)
        
        @self.bot.message_handler(commands = ['start', 'restart'])
        def start(message):
            mess = f"Привет, {message.from_user.first_name}, рад нашей встрече. Я официальный бот-помощник Liberty.\
Во мне собрана вся необходимая и полезная информация.\nЖду ваших указаний!"
            self.bot.send_message(message.chat.id, mess, reply_markup = Panel_Main.markup)

        @self.bot.message_handler(content_types = ['text', 'photo'])
        def com(message):
            self.cell[message.from_user.id].append(2)
            if (message.text == "📃Отзывы"):
                reviews(message)
            
            if (message.text == "🧮Калькулятор стоимости"):
                self.bot.send_message(message.chat.id, "Введите стоимость товара в юанях: ")
                self.bot.register_next_step_handler(message, calculation_summ)

            if (message.text == "📱Ссылки для скачивания"):
                self.bot.send_message(message.chat.id, "Выберите свою систему или вернитесь в главное меню", reply_markup = Panel_Link.markup)
                self.bot.register_next_step_handler(message, choice_platform)

            if (message.text == "🧩Как пользоваться приложениями"):
                self.bot.send_message(message.chat.id, "Выберите нужное приложение или вернитесь в главное меню", reply_markup = Panel_Video.markup)
                self.bot.register_next_step_handler(message, choice_app)

            if(message.text == "📦Инфо по доставке"):
                delivery_info(message)

            if(message.text == "❓Ответы на вопросы"):
                questions(message)
            
            if(message.text == "💬Связаться с нами"):
                if os.path.isdir(f'{message.from_user.id}'):
                    shutil.rmtree(f'{message.from_user.id}')
                    os.mkdir(f'{message.from_user.id}')
                else:
                    os.mkdir(f'{message.from_user.id}')

                workbook = Workbook()
                sheet = workbook['Sheet']
                sheet.append(['Фото', 'Размер', 'Описание', 'Цена', 'Ссылка'])
                sheet.column_dimensions['A'].width = 25
                sheet['A1'].font = font
                sheet.column_dimensions['B'].width = 25
                sheet['B1'].font = font
                sheet.column_dimensions['C'].width = 25
                sheet['C1'].font = font
                sheet.column_dimensions['D'].width = 25
                sheet['D1'].font = font
                sheet.column_dimensions['E'].width = 25
                sheet['E1'].font = font
                sheet.row_dimensions[1].height = 30
                workbook.save(f'{str(message.from_user.id)}/{str(message.from_user.id)}.xlsx')
                self.cell[message.from_user.id][0] = 2

                self.bot.send_message(message.chat.id, "Отлично, вы решили связаться с нами, но для начала необходимо заполнить таблицу!", reply_markup = Panel_Table.markup)
                self.bot.register_next_step_handler(message, connect)

        # ОТЗЫВЫ
        def reviews(message):
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton(text="Тыкай, не бойся", url = "https://t.me/Liberty_Rewiews"))
            self.bot.send_message(message.chat.id, text="Группа с отзывами", reply_markup = markup)
        
        # КАЛЬКУЛЯТОР
        def calculation_summ(message):
            if (message.text.isdigit()):
                summ = int(message.text)
                liberty_percent = summ * (liberty / 100)
                price = (summ + liberty_percent) * yan
                self.bot.send_message(message.chat.id, text=f"Курс юаня: {yan}, Наш процент: {liberty}\n\
Итоговая примерная сумма: {round(price, 2)} + {delivery_price} рублей/кг за доставку")
            
            else:
                self.bot.send_message(message.chat.id, text="Вы ввели некорректрную цену или допустили ошибку при вводе! Повторите попытку!")
                self.bot.register_next_step_handler(message, calculation_summ)
        
        # ССЫЛКИ НА СКАЧИВАНИЕ
        def choice_platform(message):
            if (message.text == "IOS"):
                self.bot.send_message(message.chat.id, text='Ссылки на приложения для IOS', reply_markup = Panel_IOS.markup)
                self.bot.send_message(message.chat.id, text="После скачивания выбранного приложения \
не забывайте обращаться к вкладке \n\"🧩Как пользоваться приложениями\"", reply_markup=Panel_Main.markup)
                
            elif (message.text == "Android"):
                self.bot.send_message(message.chat.id, text='Ссылки на приложения для Android', reply_markup = Panel_Android.markup)
                self.bot.send_message(message.chat.id, text="После скачивания выбранного приложения \
не забывайте обращаться к вкладке \n\"🧩Как пользоваться приложениями\"", reply_markup=Panel_Main.markup)
                
            elif message.text == 'Вернуться в главное меню':
                self.bot.send_message(message.chat.id, text=f"{message.from_user.first_name} вернулся в главное меню!", reply_markup=Panel_Main.markup)

            else:
                self.bot.send_message(message.chat.id, 'Вы отправили непонятную для меня команду. Выберите из предложенного списка!', reply_markup=Panel_Link.markup)
                self.bot.register_next_step_handler(message, choice_platform)
        
        # ВЫБОР ВИДЕО ДЛЯ ПРИЛОЖЕНИЯ
        def choice_app(message):
            if(message.text == "Poizon"):
                self.bot.send_video(message.chat.id, video=open("video/Poizon.mp4", "rb"), supports_streaming = True, reply_markup=Panel_Main.markup)

            elif(message.text == "TaoBao"):
                self.bot.send_video(message.chat.id, video=open("video/TaoBao.mp4", "rb"), supports_streaming = True, reply_markup=Panel_Main.markup)

            elif message.text == 'Вернуться в главное меню':
                self.bot.send_message(message.chat.id, text=f"{message.from_user.first_name} вернулся в главное меню!", reply_markup=Panel_Main.markup)
            
            else:
                self.bot.send_message(message.chat.id, 'Вы отправили непонятную для меня команду. Выберите из предложенного списка!', reply_markup=Panel_Video.markup)
                self.bot.register_next_step_handler(message, choice_app)
        
        # ИНФО О ДОСТАВКЕ
        def delivery_info(message):
            info = "1️⃣ Мы не осуществляем возврат товара, обратите на это внимание. \nСмотрите Пункт 1️⃣ \"❓Ответы на вопросы\"\n\n\
2️⃣ Прием заказов осуществляется минимум от 1000 руб. Заказы, стоимость до данного минимума не будут рассматриваться.\n\n\
3️⃣ Сумма за товар/предмет/вещь переводится сразу после оформления заявки. Это необходимо для выкупа заказа с площадки.\n\n\
4️⃣ В среднем, доставка до нас занимает 2-3 недели. Мы, в свою очередь, отправляем посылку удобным для вас способом.\n\
В большинстве случаев, доставка от нас занимает:\n\
📦Приморский край - в пределах 1 недели.\n\
📦Дальний восток - до 1,5 недели.\n\
📦Центральные регионы - около 2 недель.\n\n\
5️⃣ Если вы проживаете или временно находитесь в г. Хабаровск или г. Артем, вы можете забрать заказ лично. \
В ином случае, доставка осуществляется через 100СП, СДЕК или иную курьерскую службу по договоренности.\n\n\
6️⃣ По уточняющим вопросам обращайтесь к администраторам: @hrbdbudnf , @I_TEMKA_I"
            self.bot.send_message(message.chat.id, info)
        
        # ОТВЕТЫ НА ВОПРСЫ
        def questions(message):
            info = "1️⃣ Почему мы не можем осуществить возврат товара❓\n\
Мы являемся посредниками и работаем с КАРГО. К моменту приезда посылки в Россию, сроки возврата заканчиваются.\n\
Кроме того, стоит отметить, что посылку невозможно вернуть ввиду сложной логистики.\n\n\
2️⃣ Настоящие ли отзывы в группе❓\n\
Ответ довольно прост - если вы сомневаетесь, просто напишите покупателям и спросите их лично. Скорее всего, они будут не против поделиться опытом.\n\n\
3️⃣ Почему сумма в боте может отличаться от фактической❓\n\
Все зависит от курса юаня при обмене на бирже. Он может отличаться от того, что указан в боте.\n\
Иногда мы не успеваем его менять, не серчайте.\n\n\
4️⃣ Приходит ли оригинал с торговых площадок❓\n\
По поводу Poizon можно сказать точно - ДА. В данном маркетплейсе расположены только официальные представители компаний:\
Nike, Reebok, Adidas, TNF, Gucci, Louis Vuitton и многие другие.\
Также к каждому заказу с Poizon прикладывается сертификат, в которых расписана вся необходимая информация.\n\
По поводу TaoBao можно сказать следующее: на площадке, в большей степени, преобладает китайская атрибутика, но есть и местные качественные бренды.\
Это, буквально, Aliexpress, только изолированный от других стран.\n\
С 1688 ситуация обстоит интереснее: эта площадка представляет из себя платформу, на которой расположены страницы фабрик и Китайских поставщиков.\n\
На ней вы можете приобрести товар, непосредственно, от производителей или закупиться иной продукцией по очень низким ценам.\n\n\
5️⃣По дополнительным вопросам, вы можете обратиться к администраторам: @hrbdbudnf , @I_TEMKA_I"
            self.bot.send_message(message.chat.id, info)

        # СВЯЗЬ ( ЗАПОЛНЕНИЕ ТАБЛИЦЫ)
        def connect(message):
            if message.text == '🛒Заполнить таблицу':
                self.bot.send_message(message.chat.id, 'Отправьте фото товара в чат для добавления в таблицу!', reply_markup=Panel_Table.markup)
                self.bot.register_next_step_handler(message, add_photo)

            elif message.text == '🎬Просмотреть видео по заполнению':
                self.bot.send_message(message.chat.id, 'Извините, видео пока еще не доступно, но появится в скором времени!', reply_markup=Panel_Main.markup)
                # self.bot.send_video(message.chat.id, video=open("video/TaoBao.mp4", "rb"), supports_streaming = True, reply_markup=Panel_Main.markup)

            elif message.text == 'Вернуться в главное меню':
                self.bot.send_message(message.chat.id, text=f"{message.from_user.first_name} вернулся в главное меню!", reply_markup=Panel_Main.markup)
                shutil.rmtree(f'{message.from_user.id}')
            
            else:
                self.bot.send_message(message.chat.id, 'Вы отправили непонятную для меня команду. Выберите из предложенного списка!', reply_markup=Panel_Table.markup)
                self.bot.register_next_step_handler(message, connect)
        
        # ДОБАВЛЕНИЕ ФОТО
        def add_photo(message):
            if message.content_type == 'photo':
                photo = self.bot.get_file(message.photo[-1].file_id)
                download_photo = self.bot.download_file(photo.file_path)

                with open(f'{message.from_user.id}/{message.from_user.id}_{message.photo[-1].file_unique_id}.jpg', 'wb') as new_photo:
                    new_photo.write(download_photo)
                thing_row[message.from_user.id].append(f'{message.from_user.id}/{message.from_user.id}_{message.photo[-1].file_unique_id}.jpg')

                self.bot.send_message(message.chat.id, 'Укажите размер данного товара или отправьте \'-\', если его нет!')
                self.bot.register_next_step_handler(message, add_size)
            
            else:
                self.bot.send_message(message.chat.id, 'Необходимо отправить фото! Повторите попытку.')
                message.text = '🛒Заполнить таблицу'
                connect(message)
        
        # ДОБАВЛЕНИЕ РАЗМЕРА
        def add_size(message):
            thing_row[message.from_user.id].append(message.text)

            self.bot.send_message(message.chat.id, 'Опишите кратко товар (цвет, характеристика, кол-во и т.д.)')
            self.bot.register_next_step_handler(message, add_description)
        
        # ДОБАВЛЕНИЕ ОПИСАНИЯ
        def add_description(message):
            thing_row[message.from_user.id].append(message.text)

            self.bot.send_message(message.chat.id, 'Укажите стоимость товара в Юанях, который вы добавляете!')
            self.bot.register_next_step_handler(message, add_price)
        
        # ДОБАВЛЕНИЕ ЦЕНЫ
        def add_price(message):
            thing_row[message.from_user.id].append(message.text)
            
            self.bot.send_message(message.chat.id, 'Отправьте ссылку на товар, чтобы впоследствии легче его найти и заказать!')
            self.bot.register_next_step_handler(message, add_link)
        
        # ДОБАВЛЕНИЕ ССЫЛКИ
        def add_link(message):
            excel_file = load_workbook(f'{str(message.from_user.id)}/{str(message.from_user.id)}.xlsx')
            excel_sheet = excel_file['Sheet']

            img  = Image(thing_row[message.from_user.id][0])
            img.height = 120
            img.width = 120
            excel_sheet.add_image(img, f'A{str(self.cell[message.from_user.id][0])}')
            
            thing_row[message.from_user.id].append(message.text)
            thing_row[message.from_user.id][0] = ''
            excel_sheet.append(thing_row[message.from_user.id])
            excel_file.save(f'{str(message.from_user.id)}/{str(message.from_user.id)}.xlsx')
            thing_row[message.from_user.id].clear()
            self.cell[message.from_user.id][0] += 1

            self.bot.send_message(message.chat.id, 'Вы успешно добавили товар!\n\
Хотите добавить еще один или желаете закончить заполнение таблицы и связаться с администратором для оформления заказа?', reply_markup=Panel_Addition.markup)
            self.bot.register_next_step_handler(message, choice)
        
        # ВЫБОР: ДОБАВИТЬ ТОВАР ИЛИ ЗАКОНЧИТЬ ЗАПОЛНЕНИЕ
        def choice(message):
            if message.text == "Добавить еще один товар в заказ":
                message.text = '🛒Заполнить таблицу'
                connect(message)

            elif message.text == 'Закончить и связаться с админом':
                mess = "Ваша таблица готова! Для оформления заказа свяжитесь с администратором и отправьте ему это сообщение: @hrbdbudnf\n\n\
По техническим вопросам обращаться к: @I_TEMKA_I"

                excel_file = load_workbook(f'{str(message.from_user.id)}/{str(message.from_user.id)}.xlsx')
                excel_sheet = excel_file['Sheet']
                
                for cells in range(2, 100):
                    excel_sheet.row_dimensions[cells].height = 100
                    for cell in excel_sheet[cells]:
                        cell.alignment = alignment
                        cell.font = font
                
                excel_file.save(f'{str(message.from_user.id)}/{str(message.from_user.id)}.xlsx')
                self.bot.send_document(message.chat.id, document=open(f'{str(message.from_user.id)}/{str(message.from_user.id)}.xlsx', 'rb'), caption=mess, reply_markup=Panel_Main.markup)

                excel_sheet.delete_rows(2, 10000)
                self.cell[message.from_user.id][0] = 2
                shutil.rmtree(f'{str(message.from_user.id)}')

            else:
                self.bot.send_message(message.chat.id, 'Вы отправили непонятную для меня команду. Выберите из предложенного списка!', reply_markup=Panel_Addition.markup)
                self.bot.register_next_step_handler(message, choice)

    def run(self):  
        self.bot.polling(none_stop = True)

if __name__ == '__main__':
    Tele = TG_BOT()
    Tele.run()
